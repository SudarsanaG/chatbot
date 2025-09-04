"""
Excel Export Module for Medical Scheduling Agent
Handles data export for admin review and reporting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
from typing import Dict, List, Optional

class ExcelExporter:
    """Handles Excel export functionality for admin review"""
    
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_appointments_report(self, date_range: Optional[Dict] = None) -> str:
        """Export comprehensive appointments report"""
        try:
            # Load appointments data
            appointments_df = pd.read_excel("data/appointments.xlsx")
            
            # Filter by date range if provided
            if date_range:
                start_date = date_range.get('start_date')
                end_date = date_range.get('end_date')
                if start_date and end_date:
                    appointments_df['Date'] = pd.to_datetime(appointments_df['Date'])
                    appointments_df = appointments_df[
                        (appointments_df['Date'] >= start_date) & 
                        (appointments_df['Date'] <= end_date)
                    ]
            
            # Create filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"appointments_report_{timestamp}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Main appointments sheet
                appointments_df.to_excel(writer, sheet_name='Appointments', index=False)
                
                # Summary statistics sheet
                summary_df = self._create_summary_statistics(appointments_df)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Doctor performance sheet
                doctor_stats_df = self._create_doctor_statistics(appointments_df)
                doctor_stats_df.to_excel(writer, sheet_name='Doctor_Stats', index=False)
                
                # Patient demographics sheet
                demographics_df = self._create_demographics_report(appointments_df)
                demographics_df.to_excel(writer, sheet_name='Demographics', index=False)
            
            # Format the Excel file
            self._format_excel_file(filepath)
            
            print(f"✅ Appointments report exported to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to export appointments report: {str(e)}")
            return None
    
    def export_daily_schedule(self, date: str) -> str:
        """Export daily schedule for a specific date"""
        try:
            # Load schedules and appointments
            schedules_df = pd.read_excel("data/schedules.xlsx")
            appointments_df = pd.read_excel("data/appointments.xlsx")
            
            # Filter schedules for the specific date
            daily_schedules = schedules_df[schedules_df['Date'] == date]
            
            # Filter appointments for the specific date
            daily_appointments = appointments_df[appointments_df['Date'] == date]
            
            # Create comprehensive daily schedule
            daily_schedule_df = self._create_daily_schedule(daily_schedules, daily_appointments)
            
            # Create filename
            filename = f"daily_schedule_{date.replace('-', '')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                daily_schedule_df.to_excel(writer, sheet_name='Daily_Schedule', index=False)
                
                # Add appointment details sheet
                if not daily_appointments.empty:
                    daily_appointments.to_excel(writer, sheet_name='Appointment_Details', index=False)
            
            # Format the Excel file
            self._format_excel_file(filepath)
            
            print(f"✅ Daily schedule exported to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to export daily schedule: {str(e)}")
            return None
    
    def export_patient_database(self) -> str:
        """Export complete patient database"""
        try:
            # Load patient data
            patients_df = pd.read_csv("data/patients.csv")
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"patient_database_{timestamp}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Export to Excel with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                patients_df.to_excel(writer, sheet_name='Patients', index=False)
                
                # Add patient statistics
                stats_df = self._create_patient_statistics(patients_df)
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Format the Excel file
            self._format_excel_file(filepath)
            
            print(f"✅ Patient database exported to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to export patient database: {str(e)}")
            return None
    
    def export_revenue_report(self, date_range: Optional[Dict] = None) -> str:
        """Export revenue and billing report"""
        try:
            # Load appointments data
            appointments_df = pd.read_excel("data/appointments.xlsx")
            
            # Filter by date range if provided
            if date_range:
                start_date = date_range.get('start_date')
                end_date = date_range.get('end_date')
                if start_date and end_date:
                    appointments_df['Date'] = pd.to_datetime(appointments_df['Date'])
                    appointments_df = appointments_df[
                        (appointments_df['Date'] >= start_date) & 
                        (appointments_df['Date'] <= end_date)
                    ]
            
            # Calculate revenue (mock calculation)
            appointments_df['Revenue'] = appointments_df['Duration'] * 2  # $2 per minute
            
            # Create revenue summary
            revenue_df = self._create_revenue_summary(appointments_df)
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"revenue_report_{timestamp}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                revenue_df.to_excel(writer, sheet_name='Revenue_Summary', index=False)
                appointments_df.to_excel(writer, sheet_name='Detailed_Revenue', index=False)
            
            # Format the Excel file
            self._format_excel_file(filepath)
            
            print(f"✅ Revenue report exported to {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Failed to export revenue report: {str(e)}")
            return None
    
    def _create_summary_statistics(self, appointments_df: pd.DataFrame) -> pd.DataFrame:
        """Create summary statistics for appointments"""
        if appointments_df.empty:
            return pd.DataFrame()
        
        stats = {
            'Metric': [
                'Total Appointments',
                'New Patients',
                'Returning Patients',
                'Average Duration (minutes)',
                'Most Popular Doctor',
                'Most Popular Time Slot',
                'Appointments This Week',
                'Appointments Next Week'
            ],
            'Value': [
                len(appointments_df),
                len(appointments_df[appointments_df['PatientType'] == 'New']),
                len(appointments_df[appointments_df['PatientType'] == 'Returning']),
                appointments_df['Duration'].mean() if 'Duration' in appointments_df.columns else 0,
                appointments_df['Doctor'].mode().iloc[0] if not appointments_df['Doctor'].mode().empty else 'N/A',
                appointments_df['Time'].mode().iloc[0] if not appointments_df['Time'].mode().empty else 'N/A',
                len(appointments_df[pd.to_datetime(appointments_df['Date']).dt.isocalendar().week == datetime.now().isocalendar().week]),
                len(appointments_df[pd.to_datetime(appointments_df['Date']).dt.isocalendar().week == (datetime.now() + timedelta(weeks=1)).isocalendar().week])
            ]
        }
        
        return pd.DataFrame(stats)
    
    def _create_doctor_statistics(self, appointments_df: pd.DataFrame) -> pd.DataFrame:
        """Create doctor performance statistics"""
        if appointments_df.empty:
            return pd.DataFrame()
        
        doctor_stats = appointments_df.groupby('Doctor').agg({
            'PatientName': 'count',
            'Duration': 'sum',
            'PatientType': lambda x: (x == 'New').sum()
        }).reset_index()
        
        doctor_stats.columns = ['Doctor', 'Total_Appointments', 'Total_Duration_Minutes', 'New_Patients']
        doctor_stats['Returning_Patients'] = doctor_stats['Total_Appointments'] - doctor_stats['New_Patients']
        doctor_stats['Average_Duration'] = doctor_stats['Total_Duration_Minutes'] / doctor_stats['Total_Appointments']
        
        return doctor_stats
    
    def _create_demographics_report(self, appointments_df: pd.DataFrame) -> pd.DataFrame:
        """Create patient demographics report"""
        if appointments_df.empty:
            return pd.DataFrame()
        
        # Calculate age from DOB
        appointments_df['Age'] = pd.to_datetime(appointments_df['DOB']).apply(
            lambda x: (datetime.now() - x).days // 365
        )
        
        demographics = {
            'Age_Group': ['18-25', '26-35', '36-45', '46-55', '56-65', '65+'],
            'Count': [
                len(appointments_df[(appointments_df['Age'] >= 18) & (appointments_df['Age'] <= 25)]),
                len(appointments_df[(appointments_df['Age'] >= 26) & (appointments_df['Age'] <= 35)]),
                len(appointments_df[(appointments_df['Age'] >= 36) & (appointments_df['Age'] <= 45)]),
                len(appointments_df[(appointments_df['Age'] >= 46) & (appointments_df['Age'] <= 55)]),
                len(appointments_df[(appointments_df['Age'] >= 56) & (appointments_df['Age'] <= 65)]),
                len(appointments_df[appointments_df['Age'] > 65])
            ]
        }
        
        return pd.DataFrame(demographics)
    
    def _create_daily_schedule(self, schedules_df: pd.DataFrame, appointments_df: pd.DataFrame) -> pd.DataFrame:
        """Create comprehensive daily schedule"""
        daily_schedule = []
        
        for _, schedule_row in schedules_df.iterrows():
            doctor = schedule_row['Doctor']
            time_slot = schedule_row['Time']
            available = schedule_row['Available']
            
            # Find appointment for this slot
            appointment = appointments_df[
                (appointments_df['Doctor'] == doctor) & 
                (appointments_df['Time'] == time_slot)
            ]
            
            if not appointment.empty:
                patient_name = appointment.iloc[0]['PatientName']
                patient_type = appointment.iloc[0]['PatientType']
                duration = appointment.iloc[0]['Duration']
                status = 'Booked'
            else:
                patient_name = 'Available' if available == 'Yes' else 'Unavailable'
                patient_type = ''
                duration = 30
                status = 'Available' if available == 'Yes' else 'Unavailable'
            
            daily_schedule.append({
                'Doctor': doctor,
                'Time': time_slot,
                'Patient': patient_name,
                'Patient_Type': patient_type,
                'Duration': duration,
                'Status': status
            })
        
        return pd.DataFrame(daily_schedule)
    
    def _create_patient_statistics(self, patients_df: pd.DataFrame) -> pd.DataFrame:
        """Create patient database statistics"""
        if patients_df.empty:
            return pd.DataFrame()
        
        stats = {
            'Metric': [
                'Total Patients',
                'New Patients',
                'Returning Patients',
                'Patients with Email',
                'Patients with Phone',
                'Average Age'
            ],
            'Value': [
                len(patients_df),
                len(patients_df[patients_df['PatientType'] == 'New']),
                len(patients_df[patients_df['PatientType'] == 'Returning']),
                len(patients_df[patients_df['Email'].notna() & (patients_df['Email'] != '')]),
                len(patients_df[patients_df['Phone'].notna() & (patients_df['Phone'] != '')]),
                self._calculate_average_age(patients_df)
            ]
        }
        
        return pd.DataFrame(stats)
    
    def _create_revenue_summary(self, appointments_df: pd.DataFrame) -> pd.DataFrame:
        """Create revenue summary report"""
        if appointments_df.empty:
            return pd.DataFrame()
        
        total_revenue = appointments_df['Revenue'].sum()
        new_patient_revenue = appointments_df[appointments_df['PatientType'] == 'New']['Revenue'].sum()
        returning_patient_revenue = appointments_df[appointments_df['PatientType'] == 'Returning']['Revenue'].sum()
        
        revenue_summary = {
            'Revenue_Category': [
                'Total Revenue',
                'New Patient Revenue',
                'Returning Patient Revenue',
                'Average Revenue per Appointment',
                'Revenue per New Patient',
                'Revenue per Returning Patient'
            ],
            'Amount': [
                total_revenue,
                new_patient_revenue,
                returning_patient_revenue,
                appointments_df['Revenue'].mean(),
                new_patient_revenue / len(appointments_df[appointments_df['PatientType'] == 'New']) if len(appointments_df[appointments_df['PatientType'] == 'New']) > 0 else 0,
                returning_patient_revenue / len(appointments_df[appointments_df['PatientType'] == 'Returning']) if len(appointments_df[appointments_df['PatientType'] == 'Returning']) > 0 else 0
            ]
        }
        
        return pd.DataFrame(revenue_summary)
    
    def _calculate_average_age(self, patients_df: pd.DataFrame) -> float:
        """Calculate average age from DOB"""
        try:
            ages = pd.to_datetime(patients_df['DOB']).apply(
                lambda x: (datetime.now() - x).days // 365
            )
            return ages.mean()
        except:
            return 0
    
    def _format_excel_file(self, filepath: str):
        """Format Excel file with styling"""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to all sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Apply borders to all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = border
            
            workbook.save(filepath)
            
        except Exception as e:
            print(f"Warning: Could not format Excel file: {str(e)}")
    
    def get_export_list(self) -> List[Dict]:
        """Get list of all exported files"""
        exports = []
        
        if os.path.exists(self.export_dir):
            for filename in os.listdir(self.export_dir):
                if filename.endswith('.xlsx'):
                    filepath = os.path.join(self.export_dir, filename)
                    file_stats = os.stat(filepath)
                    
                    exports.append({
                        'filename': filename,
                        'filepath': filepath,
                        'size': file_stats.st_size,
                        'created': datetime.fromtimestamp(file_stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                        'modified': datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        return sorted(exports, key=lambda x: x['modified'], reverse=True)

# Global exporter instance
excel_exporter = ExcelExporter()

